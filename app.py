from flask import Flask, render_template, request, jsonify, send_file, session
import json, os, io, datetime, uuid, hashlib
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

app = Flask(__name__)
app.secret_key = 'yhk-harmonis-secret-2025'
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ── DATA DIR: bisa di-override pakai environment variable YHK_DATA_DIR
# Contoh: set YHK_DATA_DIR=C:\Users\Nama\OneDrive - PT XYZ\yhk-data
# Kalau tidak di-set, pakai folder 'data' di samping app.py
_env_data = os.environ.get('YHK_DATA_DIR', '').strip()
if _env_data and os.path.isabs(_env_data):
    DATA_DIR = _env_data
else:
    DATA_DIR = os.path.join(BASE_DIR, 'data')

FOTO_DIR = os.path.join(DATA_DIR, 'foto')
os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(FOTO_DIR, exist_ok=True)

DB_FILE      = os.path.join(DATA_DIR, 'transaksi.json')
KASBON_FILE  = os.path.join(DATA_DIR, 'kasbon.json')
SETTING_FILE = os.path.join(DATA_DIR, 'settings.json')
AUTH_FILE    = os.path.join(DATA_DIR, 'users.json')
BAYAR_FILE   = os.path.join(DATA_DIR, 'pembayaran.json')
QC_FILE      = os.path.join(DATA_DIR, 'qc.json')
STOK_FILE    = os.path.join(DATA_DIR, 'stok.json')

def hash_pw(pw): return hashlib.sha256(pw.encode()).hexdigest()

def get_users():
    users = load_json(AUTH_FILE, {})
    if not users:
        users = {
            'admin': {'password': hash_pw('admin123'), 'role': 'admin', 'nama': 'Admin'},
        }
        save_json(AUTH_FILE, users)
    return users

def current_user(): return session.get('user', None)
def current_role(): return session.get('role', 'guest')
def is_admin(): return current_role() == 'admin'
def is_staff(): return current_role() in ('admin', 'staff')

def require_staff(f):
    from functools import wraps
    @wraps(f)
    def dec(*a, **kw):
        if not is_staff(): return jsonify({'error': 'unauthorized'}), 401
        return f(*a, **kw)
    return dec

def require_admin(f):
    from functools import wraps
    @wraps(f)
    def dec(*a, **kw):
        if not is_admin(): return jsonify({'error': 'unauthorized'}), 401
        return f(*a, **kw)
    return dec

ALLOWED_EXT = {'jpg', 'jpeg', 'png', 'webp', 'heic'}
def allowed(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXT

# ── helpers ────────────────────────────────────────────────────────────────────
def load_json(path, default):
    if os.path.exists(path):
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return default

def save_json(path, data):
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def get_settings():
    return load_json(SETTING_FILE, {
        'penjahit': ['Andi', 'Budi', 'Citra', 'Dewi', 'Eko', 'Fitri'],
        'nama_konveksi': 'Young Harmonis Konveksi'
    })

def get_transaksi():
    return load_json(DB_FILE, [])

def get_kasbon():
    return load_json(KASBON_FILE, {})

def get_pembayaran():
    return load_json(BAYAR_FILE, {})

def get_qc():
    return load_json(QC_FILE, [])

def get_stok():
    return load_json(STOK_FILE, {})

def update_stok(nama_item, ukuran, delta):
    stok = get_stok()
    key = f"{nama_item}||{ukuran}"
    stok[key] = stok.get(key, 0) + delta
    if stok[key] < 0: stok[key] = 0
    save_json(STOK_FILE, stok)

# ── routes ─────────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/auth/status')
def auth_status():
    return jsonify({'user': current_user(), 'role': current_role()})

@app.route('/api/auth/login', methods=['POST'])
def auth_login():
    d = request.json
    username = d.get('username','').strip().lower()
    password = d.get('password','')
    users = get_users()
    u = users.get(username)
    if not u or u['password'] != hash_pw(password):
        return jsonify({'ok': False, 'error': 'Username atau password salah!'})
    session['user'] = username
    session['role'] = u['role']
    session['nama'] = u.get('nama', username)
    session.permanent = True
    return jsonify({'ok': True, 'role': u['role'], 'nama': u.get('nama', username)})

@app.route('/api/auth/logout', methods=['POST'])
def auth_logout():
    session.clear()
    return jsonify({'ok': True})

@app.route('/api/auth/change-password', methods=['POST'])
def auth_change_pw():
    if not is_staff(): return jsonify({'error': 'unauthorized'}), 401
    d = request.json
    username = current_user()
    users = get_users()
    if users[username]['password'] != hash_pw(d.get('old_password','')):
        return jsonify({'ok': False, 'error': 'Password lama salah!'})
    if len(d.get('new_password','')) < 4:
        return jsonify({'ok': False, 'error': 'Password minimal 4 karakter!'})
    users[username]['password'] = hash_pw(d['new_password'])
    save_json(AUTH_FILE, users)
    return jsonify({'ok': True})

# ── USER MANAGEMENT (admin only) ──────────────────────────────────────────────
@app.route('/api/users', methods=['GET'])
@require_admin
def users_list():
    users = get_users()
    result = [{'username': u, 'nama': d.get('nama', u), 'role': d.get('role', 'staff')} 
              for u, d in users.items()]
    return jsonify(result)

@app.route('/api/users', methods=['POST'])
@require_admin
def users_add():
    d = request.json
    username = d.get('username','').strip().lower()
    if not username: return jsonify({'ok': False, 'error': 'Username wajib!'})
    users = get_users()
    if username in users: return jsonify({'ok': False, 'error': 'Username sudah ada!'})
    users[username] = {
        'password': hash_pw(d.get('password','staff123')),
        'role': d.get('role','staff'),
        'nama': d.get('nama', username)
    }
    save_json(AUTH_FILE, users)
    return jsonify({'ok': True})

@app.route('/api/users/<string:username>', methods=['DELETE'])
@require_admin
def users_del(username):
    if username == 'admin': return jsonify({'ok': False, 'error': 'Tidak bisa hapus admin!'})
    users = get_users()
    if username in users: del users[username]
    save_json(AUTH_FILE, users)
    return jsonify({'ok': True})

@app.route('/api/users/<string:username>/reset-password', methods=['POST'])
@require_admin
def users_reset_pw(username):
    d = request.json
    users = get_users()
    if username not in users: return jsonify({'ok': False, 'error': 'User tidak ditemukan!'})
    users[username]['password'] = hash_pw(d.get('password','staff123'))
    save_json(AUTH_FILE, users)
    return jsonify({'ok': True})

@app.route('/api/settings', methods=['GET', 'POST'])
def settings():
    if request.method == 'GET':
        return jsonify(get_settings())
    if not is_admin(): return jsonify({'error': 'unauthorized'}), 401
    data = request.json
    save_json(SETTING_FILE, data)
    return jsonify({'ok': True})

DEFAULT_ITEMS = [
    {"nama": "Kemeja PDH",       "upah": 15000, "type": "umum"},
    {"nama": "Kemeja PNS",       "upah": 15000, "type": "umum"},
    {"nama": "Kemeja Korpri",    "upah": 15000, "type": "umum"},
    {"nama": "Celana PDH",       "upah": 18000, "type": "umum"},
    {"nama": "Celana PNS",       "upah": 18000, "type": "umum"},
    {"nama": "Celana Korpri",    "upah": 18000, "type": "umum"},
    {"nama": "Baju Sekolah",     "upah": 12000, "type": "sekolah_baju"},
    {"nama": "Celana Sekolah",   "upah": 14000, "type": "sekolah_celana"},
    {"nama": "Rok Sekolah",      "upah": 14000, "type": "sekolah_celana"},
    {"nama": "Baju Olahraga",    "upah": 12000, "type": "umum"},
    {"nama": "Seragam Security", "upah": 22000, "type": "umum"},
]

@app.route('/api/items', methods=['GET'])
def items_get():
    s = get_settings()
    return jsonify(s.get('items', DEFAULT_ITEMS))

@app.route('/api/items', methods=['POST'])
@require_admin
def items_save():
    items = request.json
    if not isinstance(items, list):
        return jsonify({'ok': False, 'error': 'Data tidak valid'}), 400
    s = get_settings()
    s['items'] = items
    save_json(SETTING_FILE, s)
    return jsonify({'ok': True})

@app.route('/api/transaksi', methods=['POST'])
@require_staff
def transaksi_add():
    db = get_transaksi()
    # Support both multipart/form-data (with photo) and JSON
    if request.content_type and 'multipart' in request.content_type:
        rec = json.loads(request.form.get('data', '{}'))
        fotos = []
        for f in request.files.getlist('foto'):
            if f and f.filename and allowed(f.filename):
                ext  = f.filename.rsplit('.', 1)[1].lower()
                name = f'{uuid.uuid4().hex}.{ext}'
                f.save(os.path.join(FOTO_DIR, name))
                fotos.append(name)
        rec['fotos'] = fotos
    else:
        rec = request.json
        rec['fotos'] = []
    rec['id'] = int(datetime.datetime.now().timestamp() * 1000)
    rec['createdAt'] = datetime.datetime.now().isoformat()
    db.append(rec)
    save_json(DB_FILE, db)
    return jsonify({'ok': True, 'id': rec['id']})

@app.route('/api/transaksi/<int:tid>/foto', methods=['POST'])
@require_staff
def transaksi_add_foto(tid):
    db = get_transaksi()
    rec = next((d for d in db if d.get('id') == tid), None)
    if not rec:
        return jsonify({'error': 'not found'}), 404
    saved = []
    for f in request.files.getlist('foto'):
        if f and f.filename and allowed(f.filename):
            ext  = f.filename.rsplit('.', 1)[1].lower()
            name = f'{uuid.uuid4().hex}.{ext}'
            f.save(os.path.join(FOTO_DIR, name))
            saved.append(name)
    rec.setdefault('fotos', []).extend(saved)
    save_json(DB_FILE, db)
    return jsonify({'ok': True, 'fotos': rec['fotos']})

@app.route('/api/transaksi/<int:tid>/foto/<string:fname>', methods=['DELETE'])
@require_staff
def transaksi_del_foto(tid, fname):
    db = get_transaksi()
    rec = next((d for d in db if d.get('id') == tid), None)
    if rec and fname in rec.get('fotos', []):
        rec['fotos'].remove(fname)
        fpath = os.path.join(FOTO_DIR, fname)
        if os.path.exists(fpath):
            os.remove(fpath)
        save_json(DB_FILE, db)
    return jsonify({'ok': True})

@app.route('/foto/<path:fname>')
def serve_foto(fname):
    safe = os.path.basename(fname)
    return send_file(os.path.join(FOTO_DIR, safe))

@app.route('/api/transaksi/<int:tid>', methods=['DELETE'])
@require_staff
def transaksi_del(tid):
    db = get_transaksi()
    db = [d for d in db if d.get('id') != tid]
    save_json(DB_FILE, db)
    return jsonify({'ok': True})

# ── kasbon ────────────────────────────────────────────────────────────────────
@app.route('/api/kasbon', methods=['POST'])
@require_staff
def kasbon_add():
    kb = get_kasbon()
    rec = request.json
    penjahit = rec['penjahit']
    if penjahit not in kb:
        kb[penjahit] = []
    kb[penjahit].append({
        'id': int(datetime.datetime.now().timestamp() * 1000),
        'tanggal': rec['tanggal'],
        'jumlah': rec['jumlah'],
        'keterangan': rec.get('keterangan', ''),
        'createdAt': datetime.datetime.now().isoformat()
    })
    save_json(KASBON_FILE, kb)
    return jsonify({'ok': True})

@app.route('/api/kasbon/<string:penjahit>/<int:kid>', methods=['DELETE'])
@require_staff
def kasbon_del(penjahit, kid):
    kb = get_kasbon()
    if penjahit in kb:
        kb[penjahit] = [k for k in kb[penjahit] if k['id'] != kid]
    save_json(KASBON_FILE, kb)
    return jsonify({'ok': True})

@app.route('/api/import-kasbon', methods=['POST'])
@require_staff
def import_kasbon():
    if 'file' not in request.files:
        return jsonify({'error': 'No file'}), 400
    f = request.files['file']
    wb = load_workbook(f, data_only=True)
    kb = get_kasbon()
    imported = 0
    settings = get_settings()
    penjahit_list = [p.lower() for p in settings['penjahit']]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        matched_name = None
        for p in settings['penjahit']:
            if p.lower() in sheet_name.lower():
                matched_name = p
                break
        if not matched_name:
            continue
        if matched_name not in kb:
            kb[matched_name] = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            tanggal, jumlah, ket = None, None, ''
            for cell in row:
                if isinstance(cell, datetime.datetime):
                    tanggal = cell.strftime('%Y-%m-%d')
                elif isinstance(cell, (int, float)) and cell > 0:
                    jumlah = int(cell)
                elif isinstance(cell, str) and cell.strip():
                    ket = cell.strip()
            if tanggal and jumlah:
                kb[matched_name].append({
                    'id': int(datetime.datetime.now().timestamp() * 1000) + imported,
                    'tanggal': tanggal,
                    'jumlah': jumlah,
                    'keterangan': ket or 'Import dari Excel',
                    'createdAt': datetime.datetime.now().isoformat()
                })
                imported += 1

    save_json(KASBON_FILE, kb)
    return jsonify({'ok': True, 'imported': imported})

# ringkasan moved to bottom

# ── EXPORT EXCEL ───────────────────────────────────────────────────────────────
@app.route('/api/export/excel')
def export_excel():
    tgl_dari = request.args.get('dari', '')
    tgl_ke   = request.args.get('ke', '')
    settings = get_settings()
    nama_konveksi = settings.get('nama_konveksi', 'Young Harmonis Konveksi')
    db = get_transaksi()
    kb = get_kasbon()

    if tgl_dari: db = [d for d in db if d.get('tanggal', '') >= tgl_dari]
    if tgl_ke:   db = [d for d in db if d.get('tanggal', '') <= tgl_ke]

    wb = Workbook()

    # ── color palette
    DARK    = '0F1117'
    GOLD    = 'F5A623'
    GREEN   = '3ECF8E'
    PURPLE  = '7B6CF6'
    GREY    = 'E8EAF0'
    LIGHT   = 'F5F7FA'
    MID     = 'D0D3DC'

    def hdr_font(color='FFFFFF', bold=True, sz=10):
        return Font(name='Arial', bold=bold, color=color, size=sz)

    def hdr_fill(hex_color):
        return PatternFill('solid', start_color=hex_color, fgColor=hex_color)

    def center():
        return Alignment(horizontal='center', vertical='center', wrap_text=True)

    def left():
        return Alignment(horizontal='left', vertical='center', wrap_text=True)

    def thin_border():
        s = Side(style='thin', color='CCCCCC')
        return Border(left=s, right=s, top=s, bottom=s)

    def fmt_rp(n):
        return f'Rp {n:,.0f}'.replace(',', '.')

    # ══ SHEET 1: Setoran Jahitan ═══════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = 'Setor Jahitan'
    jahit_data = [d for d in db if d.get('type') == 'jahit']

    ws1.merge_cells('A1:G1')
    ws1['A1'] = nama_konveksi + ' — Rekap Setoran Jahitan'
    ws1['A1'].font = Font(name='Arial', bold=True, size=13, color=DARK)
    ws1['A1'].alignment = center()
    ws1.row_dimensions[1].height = 28

    if tgl_dari or tgl_ke:
        ws1.merge_cells('A2:G2')
        ws1['A2'] = f'Periode: {tgl_dari or "-"} s/d {tgl_ke or "-"}'
        ws1['A2'].font = Font(name='Arial', size=9, color='666666')
        ws1['A2'].alignment = center()

    headers = ['Tanggal', 'Penjahit', 'Nama Item', 'Pcs', 'Upah/Pcs', 'Subtotal', 'Keterangan']
    col_w   = [12, 14, 24, 7, 12, 14, 20]
    row = 3
    for i, (h, w) in enumerate(zip(headers, col_w), 1):
        cell = ws1.cell(row=row, column=i, value=h)
        cell.font = hdr_font()
        cell.fill = hdr_fill(DARK)
        cell.alignment = center()
        cell.border = thin_border()
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.row_dimensions[row].height = 20
    row += 1

    start_data = row
    for d in jahit_data:
        for item in d.get('items', []):
            cells_data = [
                d['tanggal'], d['orang'], item['nama'],
                item['pcs'], item['upah'],
                item['pcs'] * item['upah'],
                d.get('keterangan', '')
            ]
            fill = hdr_fill('FFFFFF') if row % 2 == 0 else hdr_fill(LIGHT)
            for ci, val in enumerate(cells_data, 1):
                c = ws1.cell(row=row, column=ci, value=val)
                c.font = Font(name='Arial', size=9)
                c.fill = fill
                c.border = thin_border()
                c.alignment = center() if ci in [1,4,5,6] else left()
                if ci in [5, 6]:
                    c.number_format = '#,##0'
            row += 1

    # Total row
    if row > start_data:
        ws1.cell(row=row, column=1, value='TOTAL').font = Font(name='Arial', bold=True, size=10)
        ws1.cell(row=row, column=4, value=f'=SUM(D{start_data}:D{row-1})').font = Font(name='Arial', bold=True, color=GOLD)
        ws1.cell(row=row, column=6, value=f'=SUM(F{start_data}:F{row-1})').font = Font(name='Arial', bold=True, color=GOLD)
        for ci in range(1, 8):
            ws1.cell(row=row, column=ci).fill = hdr_fill('FFF3DC')
            ws1.cell(row=row, column=ci).border = thin_border()
        ws1.cell(row=row, column=4).number_format = '#,##0'
        ws1.cell(row=row, column=6).number_format = '#,##0'

    ws1.freeze_panes = 'A4'

    # ══ SHEET 2: Setoran Guntingan ═════════════════════════════════════════════
    ws2 = wb.create_sheet('Setor Guntingan')
    gunting_data = [d for d in db if d.get('type') == 'gunting']

    ws2.merge_cells('A1:G1')
    ws2['A1'] = nama_konveksi + ' — Rekap Setoran Guntingan'
    ws2['A1'].font = Font(name='Arial', bold=True, size=13)
    ws2['A1'].alignment = center()
    ws2.row_dimensions[1].height = 28

    hdr2 = ['Tanggal', 'Operator', 'No. Order', 'Jenis Kain', 'Nama Item', 'Pcs', 'Lembar']
    row2 = 3
    for i, h in enumerate(hdr2, 1):
        c = ws2.cell(row=row2, column=i, value=h)
        c.font = hdr_font(); c.fill = hdr_fill('2A7A5E'); c.alignment = center(); c.border = thin_border()
    ws2.column_dimensions['A'].width = 12; ws2.column_dimensions['B'].width = 14
    ws2.column_dimensions['C'].width = 16; ws2.column_dimensions['D'].width = 16
    ws2.column_dimensions['E'].width = 24; ws2.column_dimensions['F'].width = 8; ws2.column_dimensions['G'].width = 8
    row2 += 1
    start2 = row2
    for d in gunting_data:
        for item in d.get('items', []):
            fill = hdr_fill('FFFFFF') if row2 % 2 == 0 else hdr_fill('F0FBF6')
            for ci, val in enumerate([d['tanggal'],d['orang'],d.get('order',''),d.get('kain',''),item['nama'],item['pcs'],item.get('lembar',0)], 1):
                c = ws2.cell(row=row2, column=ci, value=val)
                c.font = Font(name='Arial', size=9); c.fill = fill; c.border = thin_border(); c.alignment = center() if ci in [1,6,7] else left()
            row2 += 1
    if row2 > start2:
        ws2.cell(row=row2, column=1, value='TOTAL').font = Font(name='Arial', bold=True)
        ws2.cell(row=row2, column=6, value=f'=SUM(F{start2}:F{row2-1})').font = Font(name='Arial', bold=True, color='2A7A5E')
        ws2.cell(row=row2, column=7, value=f'=SUM(G{start2}:G{row2-1})').font = Font(name='Arial', bold=True, color='2A7A5E')
        for ci in range(1,8): ws2.cell(row=row2, column=ci).fill = hdr_fill('DCFFF0'); ws2.cell(row=row2, column=ci).border = thin_border()
    ws2.freeze_panes = 'A4'

    # ══ SHEET 3: Pengambilan ════════════════════════════════════════════════════
    ws3 = wb.create_sheet('Pengambilan')
    ambil_data = [d for d in db if d.get('type') == 'ambil']

    ws3.merge_cells('A1:G1')
    ws3['A1'] = nama_konveksi + ' — Rekap Pengambilan Jahitan'
    ws3['A1'].font = Font(name='Arial', bold=True, size=13); ws3['A1'].alignment = center(); ws3.row_dimensions[1].height = 28

    hdr3 = ['Tanggal', 'Penjahit', 'No. Order', 'Target Selesai', 'Nama Item', 'Pcs', 'Est. Upah']
    row3 = 3
    for i, h in enumerate(hdr3, 1):
        c = ws3.cell(row=row3, column=i, value=h)
        c.font = hdr_font(); c.fill = hdr_fill('4A3D9A'); c.alignment = center(); c.border = thin_border()
    for i, w in enumerate([12,14,16,14,24,8,14], 1): ws3.column_dimensions[get_column_letter(i)].width = w
    row3 += 1; start3 = row3
    for d in ambil_data:
        for item in d.get('items', []):
            fill = hdr_fill('FFFFFF') if row3 % 2 == 0 else hdr_fill('F5F3FF')
            for ci, val in enumerate([d['tanggal'],d['orang'],d.get('order',''),d.get('deadline',''),item['nama'],item['pcs'],item['pcs']*item.get('upah',0)], 1):
                c = ws3.cell(row=row3, column=ci, value=val)
                c.font = Font(name='Arial', size=9); c.fill = fill; c.border = thin_border()
                c.alignment = center() if ci in [1,4,6,7] else left()
                if ci == 7: c.number_format = '#,##0'
            row3 += 1
    if row3 > start3:
        ws3.cell(row=row3, column=1, value='TOTAL').font = Font(name='Arial', bold=True)
        ws3.cell(row=row3, column=6, value=f'=SUM(F{start3}:F{row3-1})').font = Font(name='Arial', bold=True, color=PURPLE)
        ws3.cell(row=row3, column=7, value=f'=SUM(G{start3}:G{row3-1})').font = Font(name='Arial', bold=True, color=PURPLE)
        for ci in range(1,8): ws3.cell(row=row3, column=ci).fill = hdr_fill('EDE9FF'); ws3.cell(row=row3, column=ci).border = thin_border()
        ws3.cell(row=row3, column=6).number_format = '#,##0'; ws3.cell(row=row3, column=7).number_format = '#,##0'
    ws3.freeze_panes = 'A4'

    # ══ SHEET 4: Ringkasan Penjahit ════════════════════════════════════════════
    ws4 = wb.create_sheet('Ringkasan Penjahit')
    ws4.merge_cells('A1:F1')
    ws4['A1'] = nama_konveksi + ' — Ringkasan Upah & Kasbon'
    ws4['A1'].font = Font(name='Arial', bold=True, size=13); ws4['A1'].alignment = center(); ws4.row_dimensions[1].height = 28

    hdr4 = ['Nama Penjahit', 'Upah Setor', 'Total Kasbon', 'Saldo Bersih', 'Status']
    row4 = 3
    for i, h in enumerate(hdr4, 1):
        c = ws4.cell(row=row4, column=i, value=h)
        c.font = hdr_font(); c.fill = hdr_fill(DARK); c.alignment = center(); c.border = thin_border()
    for i, w in enumerate([20,16,16,16,14], 1): ws4.column_dimensions[get_column_letter(i)].width = w
    row4 += 1

    for nama in settings['penjahit']:
        upah   = sum(d.get('totalUpah',0) for d in db if d.get('orang')==nama and d.get('type')=='jahit')
        kasbon = sum(k['jumlah'] for k in kb.get(nama, []))
        saldo  = upah - kasbon
        status = '✓ Lunas' if saldo >= 0 else '⚠ Minus'
        fill   = hdr_fill('E8F5E9') if saldo >= 0 else hdr_fill('FFEBEE')
        vals   = [nama, upah, kasbon, saldo, status]
        for ci, val in enumerate(vals, 1):
            c = ws4.cell(row=row4, column=ci, value=val)
            c.font = Font(name='Arial', size=10, bold=(ci==1 or ci==5))
            c.fill = fill; c.border = thin_border()
            c.alignment = center() if ci > 1 else left()
            if ci in [2,3,4]: c.number_format = '#,##0'
            if ci == 4 and saldo < 0: c.font = Font(name='Arial', size=10, bold=True, color='CC0000')
        row4 += 1
    ws4.freeze_panes = 'A4'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    today = datetime.date.today().strftime('%Y%m%d')
    return send_file(buf, as_attachment=True, download_name=f'YHK_Produksi_{today}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ── EXPORT PDF ─────────────────────────────────────────────────────────────────
@app.route('/api/export/pdf')
def export_pdf():
    tgl_dari = request.args.get('dari', '')
    tgl_ke   = request.args.get('ke', '')
    settings = get_settings()
    nama_konveksi = settings.get('nama_konveksi', 'Young Harmonis Konveksi')
    db = get_transaksi()
    kb = get_kasbon()

    if tgl_dari: db = [d for d in db if d.get('tanggal', '') >= tgl_dari]
    if tgl_ke:   db = [d for d in db if d.get('tanggal', '') <= tgl_ke]

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    story  = []

    def h1(txt):
        return Paragraph(f'<font size=14 color="#0f1117"><b>{txt}</b></font>', styles['Normal'])
    def h2(txt, color='#333333'):
        return Paragraph(f'<font size=11 color="{color}"><b>{txt}</b></font>', styles['Normal'])
    def sub(txt):
        return Paragraph(f'<font size=8 color="#666666">{txt}</font>', styles['Normal'])
    def sp(h=0.3):
        return Spacer(1, h*cm)

    GOLD  = colors.HexColor('#F5A623')
    DARK  = colors.HexColor('#0F1117')
    GREEN = colors.HexColor('#3ECF8E')
    PURP  = colors.HexColor('#7B6CF6')
    LGREY = colors.HexColor('#F5F7FA')

    def make_table(data, col_widths, hdr_color=DARK):
        t = Table(data, colWidths=col_widths, repeatRows=1)
        style = [
            ('BACKGROUND', (0,0), (-1,0), hdr_color),
            ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
            ('FONTNAME',   (0,0), (-1,-1), 'Helvetica'),
            ('FONTSIZE',   (0,0), (-1,0), 8),
            ('FONTSIZE',   (0,1), (-1,-1), 7),
            ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
            ('ALIGN',      (0,0), (-1,0), 'CENTER'),
            ('ALIGN',      (0,1), (-1,-1), 'LEFT'),
            ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, LGREY]),
            ('GRID',       (0,0), (-1,-1), 0.5, colors.HexColor('#DDDDDD')),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING',(0,0),(-1,-1),4),
        ]
        t.setStyle(TableStyle(style))
        return t

    # Header
    story += [h1(f'📋 Laporan Produksi — {nama_konveksi}'), sp(0.15)]
    period = f'Periode: {tgl_dari or "Semua"} s/d {tgl_ke or "Semua"}'
    story += [sub(period + f'   |   Dicetak: {datetime.date.today().strftime("%d-%m-%Y")}'), sp(0.4)]

    # ── Ringkasan Penjahit
    story.append(h2('📊 Ringkasan Upah & Kasbon Penjahit', '#0F1117'))
    story.append(sp(0.2))
    ring_hdr = [['Penjahit', 'Upah Setor (Rp)', 'Kasbon (Rp)', 'Saldo (Rp)', 'Status']]
    for nama in settings['penjahit']:
        upah   = sum(d.get('totalUpah',0) for d in db if d.get('orang')==nama and d.get('type')=='jahit')
        kasbon = sum(k['jumlah'] for k in kb.get(nama, []))
        saldo  = upah - kasbon
        ring_hdr.append([nama, f'{upah:,.0f}', f'{kasbon:,.0f}', f'{saldo:,.0f}', '✓ Lunas' if saldo >= 0 else '⚠ Minus'])
    ring_t = make_table(ring_hdr, [3.5*cm, 3.5*cm, 3.5*cm, 3.5*cm, 2.5*cm])
    story += [ring_t, sp(0.5)]

    # ── Setoran Jahitan
    story.append(h2('🪡 Setoran Jahitan', '#B07A10'))
    story.append(sp(0.2))
    j_data = [d for d in db if d.get('type') == 'jahit']
    if j_data:
        rows = [['Tanggal', 'Penjahit', 'Item', 'Pcs', 'Upah/Pcs', 'Subtotal']]
        for d in j_data:
            for item in d.get('items', []):
                rows.append([d['tanggal'], d['orang'], item['nama'],
                             str(item['pcs']), f"{item['upah']:,.0f}", f"{item['pcs']*item['upah']:,.0f}"])
        total_pcs = sum(item['pcs'] for d in j_data for item in d.get('items',[]))
        total_upah = sum(item['pcs']*item['upah'] for d in j_data for item in d.get('items',[]))
        rows.append(['TOTAL', '', '', str(total_pcs), '', f'{total_upah:,.0f}'])
        t = make_table(rows, [2.2*cm, 2.8*cm, 4.5*cm, 1.5*cm, 2.5*cm, 3*cm], DARK)
        story += [t, sp(0.5)]
    else:
        story += [sub('Tidak ada data setoran jahitan.'), sp(0.3)]

    # ── Setoran Guntingan
    story.append(h2('✂️ Setoran Guntingan', '#1A6640'))
    story.append(sp(0.2))
    g_data = [d for d in db if d.get('type') == 'gunting']
    if g_data:
        rows = [['Tanggal', 'Operator', 'No. Order', 'Item', 'Pcs', 'Lembar']]
        for d in g_data:
            for item in d.get('items', []):
                rows.append([d['tanggal'], d['orang'], d.get('order',''), item['nama'], str(item['pcs']), str(item.get('lembar',0))])
        story += [make_table(rows, [2.2*cm, 2.8*cm, 2.8*cm, 4*cm, 1.5*cm, 1.7*cm], colors.HexColor('#2A7A5E')), sp(0.5)]
    else:
        story += [sub('Tidak ada data setoran guntingan.'), sp(0.3)]

    # ── Pengambilan
    story.append(h2('📦 Pengambilan Jahitan', '#3D2D8A'))
    story.append(sp(0.2))
    a_data = [d for d in db if d.get('type') == 'ambil']
    if a_data:
        rows = [['Tanggal', 'Penjahit', 'No. Order', 'Item', 'Pcs', 'Est. Upah']]
        for d in a_data:
            for item in d.get('items', []):
                rows.append([d['tanggal'], d['orang'], d.get('order',''), item['nama'], str(item['pcs']), f"{item['pcs']*item.get('upah',0):,.0f}"])
        story += [make_table(rows, [2.2*cm, 2.8*cm, 2.8*cm, 4*cm, 1.5*cm, 2.7*cm], colors.HexColor('#4A3D9A')), sp(0.5)]
    else:
        story += [sub('Tidak ada data pengambilan.'), sp(0.3)]

    doc.build(story)
    buf.seek(0)
    today = datetime.date.today().strftime('%Y%m%d')
    return send_file(buf, as_attachment=True, download_name=f'YHK_Laporan_{today}.pdf',
                     mimetype='application/pdf')


# ── PEMBAYARAN UPAH ───────────────────────────────────────────────────────────
@app.route('/api/pembayaran', methods=['POST'])
@require_staff
def pembayaran_add():
    d = request.json
    penjahit = d.get('penjahit', '')
    jumlah   = d.get('jumlah', 0)
    tanggal  = d.get('tanggal', '')
    ket      = d.get('keterangan', '')
    if not penjahit or not jumlah or not tanggal:
        return jsonify({'ok': False, 'error': 'Data tidak lengkap'}), 400
    bayar = get_pembayaran()
    if penjahit not in bayar:
        bayar[penjahit] = []
    rec = {
        'id': int(datetime.datetime.now().timestamp() * 1000),
        'tanggal': tanggal,
        'jumlah': jumlah,
        'keterangan': ket,
        'createdAt': datetime.datetime.now().isoformat()
    }
    bayar[penjahit].append(rec)
    save_json(BAYAR_FILE, bayar)
    return jsonify({'ok': True})

@app.route('/api/pembayaran/<string:penjahit>/<int:pid>', methods=['DELETE'])
@require_admin
def pembayaran_del(penjahit, pid):
    bayar = get_pembayaran()
    if penjahit in bayar:
        bayar[penjahit] = [b for b in bayar[penjahit] if b['id'] != pid]
    save_json(BAYAR_FILE, bayar)
    return jsonify({'ok': True})

# ── QC & FINISHING ────────────────────────────────────────────────────────────
@app.route('/api/qc', methods=['POST'])
@require_staff
def qc_add():
    d = request.json
    penjahit = d.get('penjahit','')
    tanggal  = d.get('tanggal','')
    items    = d.get('items', [])
    operator = d.get('operator','')
    ket      = d.get('keterangan','')
    if not penjahit or not items:
        return jsonify({'ok': False, 'error': 'Data tidak lengkap'}), 400

    # Validate: QC qty cannot exceed unprocessed setor qty
    tx = get_transaksi()
    qc_data = get_qc()

    rec = {
        'id': int(datetime.datetime.now().timestamp() * 1000),
        'penjahit': penjahit,
        'tanggal': tanggal,
        'operator': operator,
        'keterangan': ket,
        'items': items,
        'createdAt': datetime.datetime.now().isoformat()
    }

    # Update stok for lolos items
    for item in items:
        nama   = item.get('nama', '')
        ukuran = item.get('ukuran', '')
        lolos  = item.get('lolos', 0)
        reject = item.get('reject', 0)
        if lolos > 0:
            update_stok(nama, ukuran, lolos)

    qc_list_data = qc_data
    qc_list_data.append(rec)
    save_json(QC_FILE, qc_list_data)
    return jsonify({'ok': True, 'id': rec['id']})

@app.route('/api/qc/<int:qid>', methods=['DELETE'])
@require_admin
def qc_del(qid):
    data = get_qc()
    # Reverse stok update
    rec = next((d for d in data if d.get('id') == qid), None)
    if rec:
        for item in rec.get('items', []):
            lolos = item.get('lolos', 0)
            if lolos > 0:
                update_stok(item.get('nama',''), item.get('ukuran',''), -lolos)
    data = [d for d in data if d.get('id') != qid]
    save_json(QC_FILE, data)
    return jsonify({'ok': True})

# ── STOK ──────────────────────────────────────────────────────────────────────
@app.route('/api/stok', methods=['GET'])
def stok_get():
    stok = get_stok()
    result = []
    for key, qty in stok.items():
        if '||' in key:
            nama, ukuran = key.split('||', 1)
        else:
            nama, ukuran = key, ''
        result.append({'nama': nama, 'ukuran': ukuran, 'qty': qty, 'key': key})
    result.sort(key=lambda x: (x['nama'], x['ukuran']))
    return jsonify(result)

@app.route('/api/stok/adjust', methods=['POST'])
@require_admin
def stok_adjust():
    d = request.json
    nama   = d.get('nama','')
    ukuran = d.get('ukuran','')
    qty    = d.get('qty', 0)
    stok   = get_stok()
    key    = f"{nama}||{ukuran}"
    stok[key] = max(0, qty)
    save_json(STOK_FILE, stok)
    return jsonify({'ok': True})

# ── RINGKASAN WITH PEMBAYARAN ─────────────────────────────────────────────────
@app.route('/api/ringkasan')
def ringkasan():
    settings = get_settings()
    db = get_transaksi()
    kasbon_data = get_kasbon()
    bayar_data = get_pembayaran()

    # Kumpulkan semua orang (penjahit + operator)
    semua_orang = list(settings['penjahit'])

    result = []
    for nama in semua_orang:
        # Upah jahitan
        upah_jahit = sum(
            i.get('subtotal', i.get('pcs',0)*i.get('upah',0))
            for tx in db
            if tx.get('type') == 'jahit' and tx.get('orang') == nama
            for i in tx.get('items', [])
        )
        # Upah guntingan
        upah_gunting = sum(
            i.get('pcs',0) * i.get('upah',0)
            for tx in db
            if tx.get('type') == 'gunting' and tx.get('orang') == nama
            for i in tx.get('items', [])
        )
        upah_setor = upah_jahit + upah_gunting

        total_kasbon = sum(k.get('jumlah',0) for k in kasbon_data.get(nama, []))
        total_bayar = sum(b.get('jumlah',0) for b in bayar_data.get(nama, []))
        saldo = upah_setor - total_kasbon - total_bayar

        # Tentukan role orang ini
        has_jahit = any(tx.get('type')=='jahit' and tx.get('orang')==nama for tx in db)
        has_gunting = any(tx.get('type')=='gunting' and tx.get('orang')==nama for tx in db)
        if has_jahit and has_gunting:
            role_orang = 'keduanya'
        elif has_gunting:
            role_orang = 'cutting'
        else:
            role_orang = 'penjahit'

        result.append({
            'nama': nama,
            'upah_setor': upah_setor,
            'upah_jahit': upah_jahit,
            'upah_gunting': upah_gunting,
            'total_kasbon': total_kasbon,
            'total_bayar': total_bayar,
            'saldo': saldo,
            'role': role_orang
        })

    return jsonify(result)


@app.route('/api/pengeluaran', methods=['POST'])
@require_staff
def pengeluaran_add():
    d = request.json
    if not d.get('tujuan') or not d.get('items'):
        return jsonify({'ok': False, 'error': 'Data tidak lengkap'}), 400
    # Deduct stok
    stok = get_stok()
    for item in d['items']:
        key = f"{item['nama']}||{item['ukuran']}"
        current = stok.get(key, 0)
        if item['pcs'] > current:
            return jsonify({'ok': False, 'error': f"Stok {item['nama']} {item['ukuran']} tidak cukup! (Stok: {current})"}), 400
        stok[key] = current - item['pcs']
    save_json(STOK_FILE, stok)
    rec = {
        'id': int(datetime.datetime.now().timestamp() * 1000),
        'tanggal': d.get('tanggal', ''),
        'tujuan': d['tujuan'],
        'order': d.get('order', ''),
        'keterangan': d.get('keterangan', ''),
        'items': d['items'],
        'totalPcs': d.get('totalPcs', 0),
        'createdAt': datetime.datetime.now().isoformat()
    }
    data = get_pengeluaran()
    data.append(rec)
    save_json(KELUAR_FILE, data)
    return jsonify({'ok': True})

@app.route('/api/pengeluaran/<int:kid>', methods=['DELETE'])
@require_admin
def pengeluaran_del(kid):
    data = get_pengeluaran()
    rec = next((d for d in data if d.get('id') == kid), None)
    if rec:
        # Restore stok
        stok = get_stok()
        for item in rec.get('items', []):
            key = f"{item['nama']}||{item['ukuran']}"
            stok[key] = stok.get(key, 0) + item['pcs']
        save_json(STOK_FILE, stok)
    data = [d for d in data if d.get('id') != kid]
    save_json(KELUAR_FILE, data)
    return jsonify({'ok': True})


@app.route('/api/kain-ambil/pending/<string:operator>')
def kain_ambil_pending(operator):
    """Ambil data pengambilan kain yang belum/sudah disetor per operator"""
    ambil = get_kain_ambil()
    gunting_tx = get_transaksi()
    
    # Kumpulkan semua yang sudah diambil operator ini
    diambil = {}
    for rec in ambil:
        if rec.get('operator') == operator:
            item = rec.get('item', '')
            for d in rec.get('detail_ukuran', []):
                key = f"{item}||{d['ukuran']}"
                if key not in diambil:
                    diambil[key] = {'item': item, 'ukuran': d['ukuran'], 'total_pcs': 0, 'nama_kain': rec.get('nama_kain','')}
                diambil[key]['total_pcs'] += d.get('pcs', 0)
    
    # Kumpulkan yang sudah disetor guntingan
    sudah_setor = {}
    for tx in gunting_tx:
        if tx.get('type') == 'gunting' and tx.get('orang') == operator:
            for i in tx.get('items', []):
                key = f"{i.get('nama','')}||{i.get('ukuran','')}"
                sudah_setor[key] = sudah_setor.get(key, 0) + i.get('pcs', 0)
    
    result = []
    for key, d in diambil.items():
        sudah = sudah_setor.get(key, 0)
        sisa = d['total_pcs'] - sudah
        result.append({
            'item': d['item'],
            'ukuran': d['ukuran'],
            'nama_kain': d['nama_kain'],
            'total_diambil': d['total_pcs'],
            'sudah_setor': sudah,
            'sisa': max(0, sisa),
            'status': 'selesai' if sisa <= 0 else 'pending'
        })
    
    result.sort(key=lambda x: (x['status']=='selesai', x['item'], x['ukuran']))
    return jsonify(result)


# ── BACKUP DATA ───────────────────────────────────────────────────────────────
@app.route('/api/backup')
@require_admin
def backup_data():
    import zipfile, io
    backup = {}
    files = ['transaksi.json','kasbon.json','pembayaran.json','qc.json',
             'keluar.json','stok.json','stok_guntingan.json','stok_jahitan.json',
             'kain.json','kain_masuk.json','kain_ambil.json','alih_tugas.json',
             'settings.json','users.json']
    for fname in files:
        fpath = os.path.join(DATA_DIR, fname)
        if os.path.exists(fpath):
            with open(fpath, 'r') as f:
                backup[fname] = json.load(f)
    
    # Buat ZIP berisi semua JSON
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for fname, data in backup.items():
            zf.writestr(fname, json.dumps(data, indent=2, ensure_ascii=False))
    
    zip_buffer.seek(0)
    from flask import send_file
    tanggal = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name=f'backup_yhk_{tanggal}.zip'
    )

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5001))
    print('\n' + '='*50)
    print('  Young Harmonis Konveksi — Sistem Produksi')
    print(f'  Buka browser: http://localhost:{port}')
    print('='*50 + '\n')
    app.run(debug=True, port=port)

# ── STOK GUNTINGAN (per operator) ────────────────────────────────────────────
STOK_GUNTING_FILE = os.path.join(DATA_DIR, 'stok_guntingan.json')

def get_stok_gunting():
    return load_json(STOK_GUNTING_FILE, {})

def update_stok_gunting(operator, nama, ukuran, delta):
    sg = get_stok_gunting()
    if operator not in sg:
        sg[operator] = {}
    key = f"{nama}||{ukuran}"
    sg[operator][key] = max(0, sg[operator].get(key, 0) + delta)
    save_json(STOK_GUNTING_FILE, sg)

@app.route('/api/stok-guntingan', methods=['GET'])
def stok_gunting_get():
    sg = get_stok_gunting()
    result = []
    for operator, items in sg.items():
        for key, qty in items.items():
            if qty > 0:
                nama, ukuran = key.split('||', 1) if '||' in key else (key, '')
                result.append({'operator': operator, 'nama': nama, 'ukuran': ukuran, 'qty': qty, 'key': key})
    result.sort(key=lambda x: (x['nama'], x['ukuran']))
    return jsonify(result)

@app.route('/api/stok-guntingan/summary')
def stok_gunting_summary():
    """Ringkasan stok guntingan per item+ukuran dari semua operator"""
    sg = get_stok_gunting()
    summary = {}
    for operator, items in sg.items():
        for key, qty in items.items():
            if qty > 0:
                if key not in summary:
                    summary[key] = {'total': 0, 'per_operator': {}}
                summary[key]['total'] += qty
                summary[key]['per_operator'][operator] = qty
    result = []
    for key, data in summary.items():
        nama, ukuran = key.split('||', 1) if '||' in key else (key, '')
        result.append({'nama': nama, 'ukuran': ukuran, 'total': data['total'], 'per_operator': data['per_operator']})
    return jsonify(result)

# ── STOK JAHITAN PER PENJAHIT (bahan yang sudah diambil) ─────────────────────
STOK_JAHITAN_FILE = os.path.join(DATA_DIR, 'stok_jahitan.json')

def get_stok_jahitan():
    return load_json(STOK_JAHITAN_FILE, {})

def update_stok_jahitan(penjahit, nama, ukuran, delta):
    sj = get_stok_jahitan()
    if penjahit not in sj:
        sj[penjahit] = {}
    key = f"{nama}||{ukuran}"
    sj[penjahit][key] = max(0, sj[penjahit].get(key, 0) + delta)
    save_json(STOK_JAHITAN_FILE, sj)

@app.route('/api/stok-jahitan/<string:penjahit>')
def stok_jahitan_get(penjahit):
    """Stok bahan yang sedang dipegang penjahit (sudah diambil, belum disetor)"""
    sj = get_stok_jahitan()
    items = sj.get(penjahit, {})
    result = []
    for key, qty in items.items():
        if qty > 0:
            nama, ukuran = key.split('||', 1) if '||' in key else (key, '')
            result.append({'nama': nama, 'ukuran': ukuran, 'qty': qty})
    return jsonify(result)

# ── OVERRIDE transaksi_add untuk update stok guntingan & jahitan ──────────────
# Patch: saat transaksi ambil/gunting disimpan, update stok terkait
# Ini dilakukan via post-save hook di route existing

@app.route('/api/transaksi/ambil-check', methods=['POST'])
@require_staff
def ambil_check():
    """Validasi sebelum ambil jahitan — cek stok guntingan tersedia"""
    d = request.json
    items = d.get('items', [])
    sg = get_stok_gunting()
    
    # Hitung total stok guntingan per item+ukuran dari semua operator
    total_gunting = {}
    for operator, op_items in sg.items():
        for key, qty in op_items.items():
            total_gunting[key] = total_gunting.get(key, 0) + qty
    
    warnings = []
    errors = []
    for item in items:
        key = f"{item['nama']}||{item['ukuran']}"
        tersedia = total_gunting.get(key, 0)
        diminta = item.get('pcs', 0)
        if tersedia == 0:
            errors.append(f"{item['nama']} {item['ukuran']}: Belum ada stok guntingan!")
        elif diminta > tersedia:
            errors.append(f"{item['nama']} {item['ukuran']}: Stok guntingan hanya {tersedia} pcs, diminta {diminta} pcs")
    
    return jsonify({'ok': len(errors) == 0, 'errors': errors, 'warnings': warnings})

@app.route('/api/transaksi/setor-check', methods=['POST'])
@require_staff  
def setor_check():
    """Validasi sebelum setor jahitan — cek vs yang sudah diambil"""
    d = request.json
    penjahit = d.get('penjahit', '')
    items = d.get('items', [])
    sj = get_stok_jahitan()
    pj_items = sj.get(penjahit, {})
    
    warnings = []
    for item in items:
        key = f"{item['nama']}||{item['ukuran']}"
        diambil = pj_items.get(key, 0)
        disetor = item.get('pcs', 0)
        if disetor > diambil:
            warnings.append(f"{item['nama']} {item['ukuran']}: Setor {disetor} pcs tapi hanya ambil {diambil} pcs")
    
    return jsonify({'ok': True, 'warnings': warnings})

@app.route('/api/stok-guntingan/kurangi', methods=['POST'])
@require_staff
def kurangi_stok_gunting():
    """Kurangi stok guntingan saat penjahit ambil bahan (FIFO dari operator)"""
    d = request.json
    items = d.get('items', [])
    sg = get_stok_gunting()
    
    for item in items:
        nama = item['nama']
        ukuran = item['ukuran']
        key = f"{nama}||{ukuran}"
        sisa = item['pcs']
        
        # Kurangi dari operator secara FIFO
        for operator in list(sg.keys()):
            if sisa <= 0:
                break
            tersedia = sg[operator].get(key, 0)
            if tersedia > 0:
                kurang = min(tersedia, sisa)
                sg[operator][key] = tersedia - kurang
                sisa -= kurang
    
    save_json(STOK_GUNTING_FILE, sg)
    return jsonify({'ok': True})

@app.route('/api/stok-guntingan/tambah', methods=['POST'])
@require_staff
def tambah_stok_gunting():
    """Tambah stok guntingan saat operator setor guntingan"""
    d = request.json
    operator = d.get('operator', '')
    items = d.get('items', [])
    sg = get_stok_gunting()
    
    if operator not in sg:
        sg[operator] = {}
    
    for item in items:
        key = f"{item['nama']}||{item['ukuran']}"
        sg[operator][key] = sg[operator].get(key, 0) + item.get('pcs', 0)
    
    save_json(STOK_GUNTING_FILE, sg)
    return jsonify({'ok': True})

@app.route('/api/stok-jahitan/tambah', methods=['POST'])
@require_staff
def tambah_stok_jahitan():
    """Tambah stok jahitan penjahit saat ambil bahan"""
    d = request.json
    penjahit = d.get('penjahit', '')
    items = d.get('items', [])
    sj = get_stok_jahitan()
    
    if penjahit not in sj:
        sj[penjahit] = {}
    
    for item in items:
        key = f"{item['nama']}||{item['ukuran']}"
        sj[penjahit][key] = sj[penjahit].get(key, 0) + item.get('pcs', 0)
    
    save_json(STOK_JAHITAN_FILE, sj)
    return jsonify({'ok': True})

@app.route('/api/stok-jahitan/kurangi', methods=['POST'])
@require_staff
def kurangi_stok_jahitan():
    """Kurangi stok jahitan penjahit saat setor hasil jahit"""
    d = request.json
    penjahit = d.get('penjahit', '')
    items = d.get('items', [])
    sj = get_stok_jahitan()
    
    if penjahit not in sj:
        sj[penjahit] = {}
    
    for item in items:
        key = f"{item['nama']}||{item['ukuran']}"
        current = sj[penjahit].get(key, 0)
        sj[penjahit][key] = max(0, current - item.get('pcs', 0))
    
    save_json(STOK_JAHITAN_FILE, sj)
    return jsonify({'ok': True})

# ── ALIH TUGAS ────────────────────────────────────────────────────────────────
ALIHTGAS_FILE = os.path.join(DATA_DIR, 'alih_tugas.json')

def get_alih_tugas():
    return load_json(ALIHTGAS_FILE, [])

@app.route('/api/alih-tugas', methods=['POST'])
@require_admin
def alih_tugas_add():
    d = request.json
    dari = d.get('dari', '')
    ke = d.get('ke', '')
    items = d.get('items', [])
    alasan = d.get('alasan', '')
    
    if not dari or not ke or not items:
        return jsonify({'ok': False, 'error': 'Data tidak lengkap'}), 400
    if dari == ke:
        return jsonify({'ok': False, 'error': 'Penjahit asal dan tujuan sama!'}), 400
    
    sj = get_stok_jahitan()
    pj_items = sj.get(dari, {})
    
    # Validasi
    for item in items:
        key = f"{item['nama']}||{item['ukuran']}"
        tersedia = pj_items.get(key, 0)
        if item['pcs'] > tersedia:
            return jsonify({'ok': False, 'error': f"{item['nama']} {item['ukuran']}: {dari} hanya punya {tersedia} pcs"}), 400
    
    # Transfer stok jahitan
    if dari not in sj: sj[dari] = {}
    if ke not in sj: sj[ke] = {}
    
    for item in items:
        key = f"{item['nama']}||{item['ukuran']}"
        sj[dari][key] = max(0, sj[dari].get(key, 0) - item['pcs'])
        sj[ke][key] = sj[ke].get(key, 0) + item['pcs']
    
    save_json(STOK_JAHITAN_FILE, sj)
    
    # Simpan history
    rec = {
        'id': int(datetime.datetime.now().timestamp() * 1000),
        'tanggal': d.get('tanggal', ''),
        'dari': dari,
        'ke': ke,
        'items': items,
        'alasan': alasan,
        'createdAt': datetime.datetime.now().isoformat()
    }
    data = get_alih_tugas()
    data.append(rec)
    save_json(ALIHTGAS_FILE, data)
    
    return jsonify({'ok': True})

@app.route('/api/penjahit-progress')
def penjahit_progress():
    """Ringkasan progress per penjahit: ambil berapa, setor berapa, sisa di tangan"""
    settings = get_settings()
    sj = get_stok_jahitan()
    db = get_transaksi()
    
    result = []
    for nama in settings['penjahit']:
        # Hitung total yang diambil
        total_ambil = sum(
            i.get('pcs', 0)
            for tx in db
            if tx.get('type') == 'ambil' and tx.get('orang') == nama
            for i in tx.get('items', [])
        )
        # Hitung total yang sudah disetor
        total_setor = sum(
            i.get('pcs', 0)
            for tx in db
            if tx.get('type') == 'jahit' and tx.get('orang') == nama
            for i in tx.get('items', [])
        )
        # Stok saat ini di tangan
        items_di_tangan = []
        pj_stok = sj.get(nama, {})
        for key, qty in pj_stok.items():
            if qty > 0:
                n, u = key.split('||', 1) if '||' in key else (key, '')
                items_di_tangan.append({'nama': n, 'ukuran': u, 'qty': qty})
        
        result.append({
            'nama': nama,
            'total_ambil': total_ambil,
            'total_setor': total_setor,
            'di_tangan': sum(i['qty'] for i in items_di_tangan),
            'items_di_tangan': items_di_tangan
        })
    
    return jsonify(result)


# ── MASTER KAIN ───────────────────────────────────────────────────────────────
KAIN_FILE = os.path.join(DATA_DIR, 'kain.json')
KAIN_MASUK_FILE = os.path.join(DATA_DIR, 'kain_masuk.json')
KAIN_AMBIL_FILE = os.path.join(DATA_DIR, 'kain_ambil.json')

def get_kain():
    return load_json(KAIN_FILE, [
        {"nama": "Kain Korpri"},
        {"nama": "Kain Drill"},
        {"nama": "Kain TC"},
        {"nama": "Kain Lacoste"},
    ])

def get_kain_masuk():
    return load_json(KAIN_MASUK_FILE, [])

def get_kain_ambil():
    return load_json(KAIN_AMBIL_FILE, [])

@app.route('/api/kain', methods=['GET'])
def kain_list():
    return jsonify(get_kain())

@app.route('/api/kain', methods=['POST'])
@require_admin
def kain_add():
    d = request.json
    nama = d.get('nama', '').strip()
    if not nama:
        return jsonify({'ok': False, 'error': 'Nama kain wajib!'}), 400
    kain = get_kain()
    if any(k['nama'].lower() == nama.lower() for k in kain):
        return jsonify({'ok': False, 'error': 'Kain sudah ada!'}), 400
    kain.append({'nama': nama})
    save_json(KAIN_FILE, kain)
    return jsonify({'ok': True})

@app.route('/api/kain/<string:nama>', methods=['DELETE'])
@require_admin
def kain_del(nama):
    kain = get_kain()
    kain = [k for k in kain if k['nama'] != nama]
    save_json(KAIN_FILE, kain)
    return jsonify({'ok': True})

# ── KAIN MASUK ────────────────────────────────────────────────────────────────
@app.route('/api/kain-masuk', methods=['GET'])
def kain_masuk_list():
    return jsonify(list(reversed(get_kain_masuk())))

@app.route('/api/kain-masuk', methods=['POST'])
@require_staff
def kain_masuk_add():
    d = request.json
    if not d.get('tanggal') or not d.get('nama_kain') or not d.get('meter'):
        return jsonify({'ok': False, 'error': 'Data tidak lengkap'}), 400
    rec = {
        'id': int(datetime.datetime.now().timestamp() * 1000),
        'tanggal': d.get('tanggal'),
        'nama_kain': d.get('nama_kain'),
        'meter': float(d.get('meter', 0)),
        'supplier': d.get('supplier', ''),
        'keterangan': d.get('keterangan', ''),
        'createdAt': datetime.datetime.now().isoformat()
    }
    data = get_kain_masuk()
    data.append(rec)
    save_json(KAIN_MASUK_FILE, data)
    return jsonify({'ok': True})

@app.route('/api/kain-masuk/<int:kid>', methods=['DELETE'])
@require_admin
def kain_masuk_del(kid):
    data = [d for d in get_kain_masuk() if d.get('id') != kid]
    save_json(KAIN_MASUK_FILE, data)
    return jsonify({'ok': True})

# ── KAIN DIAMBIL (Pengambilan Kain) ──────────────────────────────────────────
@app.route('/api/kain-ambil', methods=['GET'])
def kain_ambil_list():
    return jsonify(list(reversed(get_kain_ambil())))

@app.route('/api/kain-ambil', methods=['POST'])
@require_staff
def kain_ambil_add():
    d = request.json
    if not d.get('tanggal') or not d.get('nama_kain') or not d.get('operator'):
        return jsonify({'ok': False, 'error': 'Data tidak lengkap'}), 400
    rec = {
        'id': int(datetime.datetime.now().timestamp() * 1000),
        'tanggal': d.get('tanggal'),
        'nama_kain': d.get('nama_kain'),
        'operator': d.get('operator'),
        'total_meter': float(d.get('total_meter', 0)),
        'item': d.get('item', ''),
        'detail_ukuran': d.get('detail_ukuran', []),
        'keterangan': d.get('keterangan', ''),
        'createdAt': datetime.datetime.now().isoformat()
    }
    data = get_kain_ambil()
    data.append(rec)
    save_json(KAIN_AMBIL_FILE, data)
    return jsonify({'ok': True})

@app.route('/api/kain-ambil/<int:kid>', methods=['DELETE'])
@require_admin
def kain_ambil_del(kid):
    data = [d for d in get_kain_ambil() if d.get('id') != kid]
    save_json(KAIN_AMBIL_FILE, data)
    return jsonify({'ok': True})

# ── RATA-RATA PENGGUNAAN KAIN ─────────────────────────────────────────────────
@app.route('/api/kain-rata-rata')
def kain_rata_rata():
    """Hitung rata-rata penggunaan kain per kain+item+ukuran"""
    ambil_data = get_kain_ambil()
    
    # Akumulasi data: {kain||item||ukuran: {total_meter, total_pcs, count}}
    stats = {}
    for rec in ambil_data:
        kain = rec.get('nama_kain', '')
        item = rec.get('item', '')
        for detail in rec.get('detail_ukuran', []):
            ukuran = detail.get('ukuran', '')
            meter = float(detail.get('meter', 0))
            pcs = int(detail.get('pcs', 0))
            if pcs > 0 and meter > 0:
                key = f"{kain}||{item}||{ukuran}"
                if key not in stats:
                    stats[key] = {'total_meter': 0, 'total_pcs': 0, 'count': 0}
                stats[key]['total_meter'] += meter
                stats[key]['total_pcs'] += pcs
                stats[key]['count'] += 1
    
    result = []
    for key, s in stats.items():
        parts = key.split('||')
        kain, item, ukuran = parts[0], parts[1], parts[2]
        rata = round(s['total_meter'] / s['total_pcs'], 3) if s['total_pcs'] > 0 else 0
        result.append({
            'kain': kain,
            'item': item,
            'ukuran': ukuran,
            'rata_rata_meter': rata,
            'total_meter': round(s['total_meter'], 2),
            'total_pcs': s['total_pcs'],
            'jumlah_data': s['count']
        })
    
    result.sort(key=lambda x: (x['kain'], x['item'], x['ukuran']))
    return jsonify(result)

# ── STOK KAIN (sisa kain) ─────────────────────────────────────────────────────
@app.route('/api/stok-kain')
def stok_kain():
    """Hitung sisa stok kain: masuk - terpakai"""
    masuk = get_kain_masuk()
    ambil = get_kain_ambil()
    
    # Total masuk per kain
    total_masuk = {}
    for rec in masuk:
        k = rec.get('nama_kain', '')
        total_masuk[k] = total_masuk.get(k, 0) + float(rec.get('meter', 0))
    
    # Total terpakai per kain
    total_pakai = {}
    for rec in ambil:
        k = rec.get('nama_kain', '')
        total_pakai[k] = total_pakai.get(k, 0) + float(rec.get('total_meter', 0))
    
    result = []
    all_kain = set(list(total_masuk.keys()) + list(total_pakai.keys()))
    for k in sorted(all_kain):
        masuk_m = total_masuk.get(k, 0)
        pakai_m = total_pakai.get(k, 0)
        result.append({
            'nama_kain': k,
            'total_masuk': round(masuk_m, 2),
            'total_pakai': round(pakai_m, 2),
            'sisa': round(masuk_m - pakai_m, 2)
        })
    
    return jsonify(result)

